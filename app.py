import html
import io
import re
import time
import unicodedata
from dataclasses import dataclass
from difflib import SequenceMatcher
from typing import Dict, List, Optional, Tuple
from urllib.parse import parse_qs, quote, unquote, urlparse
from urllib.request import Request, urlopen

import cv2
import fitz  # PyMuPDF
import numpy as np
import pandas as pd
import streamlit as st
from PIL import Image, ImageDraw
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from pyzbar.pyzbar import decode as pyzbar_decode
except Exception:
    pyzbar_decode = None


REQUIRED_COLUMNS = ["페이지", "카드번호", "제목", "유튜브링크"]


@dataclass
class PdfLink:
    page: int
    x0: float
    y0: float
    x1: float
    y1: float
    url: str
    kind: str
    target_page: Optional[int] = None


@dataclass
class PdfQr:
    page: int
    x0: float
    y0: float
    x1: float
    y1: float
    url: str


# -----------------------------
# Basic parsing / comparison
# -----------------------------
def read_csv_flexible(file_bytes: bytes) -> pd.DataFrame:
    """CSV 인코딩이 UTF-8, UTF-8 BOM, CP949여도 읽을 수 있게 처리."""
    last_error = None
    for enc in ["utf-8-sig", "utf-8", "cp949", "euc-kr"]:
        try:
            return pd.read_csv(io.BytesIO(file_bytes), encoding=enc)
        except Exception as exc:
            last_error = exc
    raise ValueError(f"CSV를 읽지 못했습니다. 인코딩 또는 파일 형식을 확인해 주세요. 원인: {last_error}")


def normalize_text(value: str) -> str:
    value = "" if value is None else str(value)
    value = unicodedata.normalize("NFKC", value)
    value = value.replace("’", "'").replace("–", "-").replace("—", "-")
    value = re.sub(r"\s+", "", value)
    return value.lower().strip()


def normalize_title_for_similarity(value: str) -> str:
    value = "" if value is None else str(value)
    value = html.unescape(value)
    value = unicodedata.normalize("NFKC", value)
    value = value.lower()
    # YouTube title often includes channel/series labels. They should not dominate similarity.
    noise_patterns = [
        r"\bkia\s+global\s+how\s*to\b",
        r"\bkia\s+how\s*to\b",
        r"\bhow\s*to\b",
        r"\bofficial\b",
        r"\bvideo\b",
        r"\bguide\b",
        r"\bthe\s+kia\s+k4\b",
    ]
    for pat in noise_patterns:
        value = re.sub(pat, " ", value, flags=re.IGNORECASE)
    value = value.replace("&amp;", "&")
    value = re.sub(r"[^a-z0-9가-힣]+", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def title_similarity(a: str, b: str) -> float:
    a_norm = normalize_title_for_similarity(a)
    b_norm = normalize_title_for_similarity(b)
    if not a_norm or not b_norm:
        return 0.0
    if a_norm == b_norm:
        return 1.0
    seq = SequenceMatcher(None, a_norm, b_norm).ratio()
    a_tokens = set(a_norm.split())
    b_tokens = set(b_norm.split())
    token_score = len(a_tokens & b_tokens) / max(len(a_tokens | b_tokens), 1)
    contain_score = 0.0
    if a_norm in b_norm or b_norm in a_norm:
        contain_score = min(len(a_norm), len(b_norm)) / max(len(a_norm), len(b_norm))
    return round(max(seq, token_score, contain_score), 3)


def normalize_url(value: str) -> str:
    if not value:
        return ""
    return unquote(str(value).strip())


def parse_url_parts(url: str) -> Dict[str, Optional[str]]:
    """YouTube URL은 영상 ID / playlist ID / 채널 핸들을 분리해서 비교."""
    url = normalize_url(url)
    out = {
        "raw": url,
        "domain": None,
        "type": "other",
        "video_id": None,
        "playlist_id": None,
        "channel": None,
        "index": None,
    }
    if not url:
        return out

    try:
        p = urlparse(url)
        domain = p.netloc.lower().replace("www.", "")
        out["domain"] = domain
        qs = parse_qs(p.query)
        out["index"] = (qs.get("index") or [None])[0]
        out["playlist_id"] = (qs.get("list") or [None])[0]

        if "youtube.com" in domain:
            if p.path == "/watch":
                out["type"] = "video"
                out["video_id"] = (qs.get("v") or [None])[0]
            elif p.path.startswith("/shorts/"):
                out["type"] = "video"
                out["video_id"] = p.path.split("/shorts/", 1)[1].split("/")[0]
            elif p.path.startswith("/playlist"):
                out["type"] = "playlist"
            elif p.path.startswith("/@"):
                out["type"] = "channel"
                out["channel"] = p.path.strip("/")
            elif p.path.startswith("/channel/") or p.path.startswith("/c/") or p.path.startswith("/user/"):
                out["type"] = "channel"
                out["channel"] = p.path.strip("/")
        elif "youtu.be" in domain:
            out["type"] = "video"
            out["video_id"] = p.path.strip("/").split("/")[0]
    except Exception:
        pass
    return out


def canonical_youtube_url(url: str) -> str:
    parts = parse_url_parts(url)
    if parts.get("video_id"):
        return f"https://www.youtube.com/watch?v={parts['video_id']}"
    return normalize_url(url)


def compare_urls(expected: str, actual: str) -> Tuple[str, str]:
    if not actual:
        return "MISSING", "URL 없음"

    e = parse_url_parts(expected)
    a = parse_url_parts(actual)

    if normalize_url(expected) == normalize_url(actual):
        return "EXACT", "URL 전체 일치"

    if e["video_id"] and a["video_id"] and e["video_id"] == a["video_id"]:
        note = "영상 ID 일치"
        if e.get("index") != a.get("index"):
            note += f" / index 다름: CSV {e.get('index')} vs PDF {a.get('index')}"
        return "VIDEO_ID_MATCH", note

    if e["playlist_id"] and a["playlist_id"] and e["playlist_id"] == a["playlist_id"]:
        return "PLAYLIST_MATCH", "플레이리스트 ID 일치"

    if e["channel"] and a["channel"] and normalize_text(e["channel"]) == normalize_text(a["channel"]):
        return "CHANNEL_MATCH", "채널 일치"

    return "MISMATCH", "기대 URL과 다름"


def compare_two_actual_urls(url_a: str, url_b: str) -> Tuple[str, str]:
    if not url_a and not url_b:
        return "MISSING", "QR과 하이퍼링크가 모두 없음"
    if not url_a:
        return "MISSING_QR", "QR URL 없음"
    if not url_b:
        return "MISSING_LINK", "하이퍼링크 URL 없음"
    return compare_urls(url_a, url_b)


def is_external_video_url(url: str) -> bool:
    if not url:
        return False
    parts = parse_url_parts(url)
    return parts["type"] in {"video", "playlist", "channel"}


# -----------------------------
# YouTube oEmbed title lookup
# -----------------------------
@st.cache_data(show_spinner=False, ttl=60 * 60 * 24)
def fetch_youtube_title_cached(url: str) -> Tuple[str, str]:
    """Return (title, error). Uses YouTube oEmbed, no API key required."""
    target = canonical_youtube_url(url)
    if not target:
        return "", "URL 없음"
    parts = parse_url_parts(target)
    if parts.get("type") != "video" or not parts.get("video_id"):
        return "", "영상 URL이 아니라 실제 제목 조회를 생략함"

    api = f"https://www.youtube.com/oembed?url={quote(target, safe=':/?=&')}&format=json"
    try:
        req = Request(api, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(req, timeout=8) as resp:
            raw = resp.read().decode("utf-8", errors="replace")
        m = re.search(r'"title"\s*:\s*"(.*?)"', raw)
        if not m:
            return "", "YouTube 제목 응답을 해석하지 못함"
        # crude but enough for JSON string title; avoid adding dependency
        title = bytes(m.group(1), "utf-8").decode("unicode_escape")
        title = html.unescape(title).strip()
        return title, ""
    except Exception as exc:
        return "", f"YouTube 제목 조회 실패: {exc}"


def judge_youtube_title(
    youtube_title: str,
    csv_title: str,
    pdf_title: str,
    ok_threshold: float,
    check_threshold: float,
) -> Tuple[str, float, float, str]:
    if not youtube_title:
        return "SKIP", 0.0, 0.0, "유튜브 제목 없음"
    csv_score = title_similarity(youtube_title, csv_title)
    pdf_score = title_similarity(youtube_title, pdf_title) if pdf_title else 0.0
    best = max(csv_score, pdf_score)
    note = f"CSV 유사도 {csv_score:.2f} / PDF 유사도 {pdf_score:.2f}"
    if best >= ok_threshold:
        return "OK", csv_score, pdf_score, note
    if best >= check_threshold:
        return "CHECK", csv_score, pdf_score, note
    return "MISMATCH", csv_score, pdf_score, note


# -----------------------------
# PDF extraction
# -----------------------------
def load_pdf(pdf_bytes: bytes) -> fitz.Document:
    return fitz.open(stream=pdf_bytes, filetype="pdf")


def extract_links(doc: fitz.Document) -> List[PdfLink]:
    links: List[PdfLink] = []
    for page_index, page in enumerate(doc, start=1):
        for link in page.get_links():
            rect = link.get("from")
            if not rect:
                continue
            uri = link.get("uri") or ""
            kind = "external" if uri else "internal"
            target_page = None
            if not uri and link.get("page") is not None:
                target_page = int(link.get("page")) + 1
            links.append(
                PdfLink(
                    page=page_index,
                    x0=float(rect.x0),
                    y0=float(rect.y0),
                    x1=float(rect.x1),
                    y1=float(rect.y1),
                    url=uri,
                    kind=kind,
                    target_page=target_page,
                )
            )
    return links


def extract_qrs(doc: fitz.Document, zoom: float = 2.0, expected_counts: Optional[Dict[int, int]] = None) -> List[PdfQr]:
    qrs: List[PdfQr] = []
    detector = cv2.QRCodeDetector()
    matrix = fitz.Matrix(zoom, zoom)
    seen = set()

    def add_qr(page_index: int, data: str, xs, ys):
        data = str(data).strip()
        if not data:
            return
        x0, y0, x1, y1 = float(min(xs)), float(min(ys)), float(max(xs)), float(max(ys))
        key = (page_index, data, round(x0, 1), round(y0, 1))
        if key in seen:
            return
        seen.add(key)
        qrs.append(PdfQr(page=page_index, x0=x0, y0=y0, x1=x1, y1=y1, url=data))

    for page_index, page in enumerate(doc, start=1):
        pix = page.get_pixmap(matrix=matrix, alpha=False)
        img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.h, pix.w, pix.n)
        if pix.n == 4:
            img = cv2.cvtColor(img, cv2.COLOR_RGBA2RGB)
        bgr = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)

        ok, decoded_info, points, _ = detector.detectAndDecodeMulti(bgr)
        if ok and points is not None:
            for data, pts in zip(decoded_info, points):
                xs = pts[:, 0] / zoom
                ys = pts[:, 1] / zoom
                add_qr(page_index, data, xs, ys)

        current_page_count = sum(1 for item in qrs if item.page == page_index)
        need_extra_scan = False
        if expected_counts and page_index in expected_counts:
            need_extra_scan = current_page_count < int(expected_counts.get(page_index, 0))

        if pyzbar_decode is not None and need_extra_scan:
            try:
                pil_img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                for item in pyzbar_decode(pil_img):
                    data = item.data.decode("utf-8", errors="replace")
                    rect = item.rect
                    xs = [rect.left / zoom, (rect.left + rect.width) / zoom]
                    ys = [rect.top / zoom, (rect.top + rect.height) / zoom]
                    add_qr(page_index, data, xs, ys)
            except Exception:
                pass

    qrs.sort(key=lambda r: (r.page, r.x0, r.y0))
    return qrs


def page_text_spans(page: fitz.Page) -> List[Dict]:
    spans = []
    data = page.get_text("dict")
    for block in data.get("blocks", []):
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            line_text = " ".join(span.get("text", "").strip() for span in line.get("spans", []) if span.get("text", "").strip())
            if not line_text:
                continue
            xs0, ys0, xs1, ys1 = [], [], [], []
            for span in line.get("spans", []):
                text = span.get("text", "").strip()
                if not text:
                    continue
                x0, y0, x1, y1 = span.get("bbox")
                xs0.append(x0); ys0.append(y0); xs1.append(x1); ys1.append(y1)
            if xs0:
                spans.append({"text": line_text, "x0": min(xs0), "y0": min(ys0), "x1": max(xs1), "y1": max(ys1)})
    return spans


def extract_title_text_for_region(page: fitz.Page, x0: float, x1: float, y_min: float, y_max: float) -> str:
    spans = page_text_spans(page)
    selected = []
    for sp in spans:
        cx = (sp["x0"] + sp["x1"]) / 2
        cy = (sp["y0"] + sp["y1"]) / 2
        if x0 <= cx <= x1 and y_min <= cy <= y_max:
            selected.append(sp)
    selected.sort(key=lambda s: (round(s["y0"] / 4) * 4, s["x0"]))
    return " ".join(s["text"] for s in selected).strip()


def find_best_title_in_card(page: fitz.Page, expected_title: str, x0: float, x1: float, y_min: float, y_max: float) -> Tuple[str, float]:
    spans = page_text_spans(page)
    candidates = []
    # Wider search inside the card, but still avoids footer/nav areas.
    for sp in spans:
        cx = (sp["x0"] + sp["x1"]) / 2
        cy = (sp["y0"] + sp["y1"]) / 2
        if x0 <= cx <= x1 and 100 <= cy <= min(page.rect.height - 80, 680):
            text = sp["text"].strip()
            if not text:
                continue
            # Skip repeated helper/footer text.
            skip_tokens = ["click to play", "scan the qr", "youtube video", "kia how to package", "introduction category"]
            if any(tok in text.lower() for tok in skip_tokens):
                continue
            score = title_similarity(text, expected_title)
            if score > 0:
                candidates.append((score, text, sp["y0"]))
    if candidates:
        candidates.sort(key=lambda item: (-item[0], item[2]))
        return candidates[0][1], candidates[0][0]
    fallback = extract_title_text_for_region(page, x0, x1, y_min, y_max)
    return fallback, title_similarity(fallback, expected_title)


def page_has_text(page: fitz.Page, expected_title: str) -> bool:
    text = page.get_text("text") or ""
    return normalize_text(expected_title) in normalize_text(text)


def group_by_page(items):
    grouped: Dict[int, List] = {}
    for item in items:
        grouped.setdefault(item.page, []).append(item)
    for page in grouped:
        grouped[page].sort(key=lambda r: (r.x0, r.y0))
    return grouped


def row_card_rect(page: fitz.Page, card_count: int, card_no: int, link: Optional[PdfLink], qr: Optional[PdfQr]) -> Tuple[float, float, float, float]:
    if link:
        return link.x0, link.y0, link.x1, link.y1
    if qr:
        width = page.rect.width
        approx_width = width / max(card_count, 1)
        cx = (qr.x0 + qr.x1) / 2
        return max(0, cx - approx_width / 2), 0, min(width, cx + approx_width / 2), page.rect.height

    width = page.rect.width
    margin = width * 0.04
    usable = width - margin * 2
    card_w = usable / max(card_count, 1)
    x0 = margin + (card_no - 1) * card_w
    x1 = margin + card_no * card_w
    return x0, 0, x1, page.rect.height


# -----------------------------
# Validation
# -----------------------------
def action_memo(issues: List[str], detail_notes: List[str]) -> str:
    if not issues and not detail_notes:
        return "수정 없음"
    messages = []
    if "TITLE_MISMATCH" in issues:
        messages.append("PDF 카드 제목을 CSV 제목과 맞춰 확인하세요.")
    if "TITLE_CHECK" in issues:
        messages.append("제목 위치/카드 매칭이 애매합니다. 페이지 미리보기에서 카드 위치를 확인하세요.")
    if "MISSING_QR" in issues:
        messages.append("QR이 없거나 인식되지 않았습니다. QR 이미지 선명도와 위치를 확인하세요.")
    if "QR_MISMATCH" in issues:
        messages.append("QR에 들어간 URL이 CSV 기준 링크와 다릅니다.")
    if "MISSING_LINK" in issues:
        messages.append("카드/Click to play 영역에 하이퍼링크를 추가하세요.")
    if "LINK_MISMATCH" in issues:
        messages.append("카드 하이퍼링크 URL이 CSV 기준 링크와 다릅니다.")
    if "QR_LINK_MISMATCH" in issues:
        messages.append("QR URL과 카드 하이퍼링크 URL이 서로 다릅니다.")
    if "YOUTUBE_TITLE_MISMATCH" in issues:
        messages.append("연결된 유튜브 실제 제목이 CSV/PDF 제목과 크게 다릅니다. 링크가 맞는지 확인하세요.")
    if "YOUTUBE_TITLE_CHECK" in issues:
        messages.append("유튜브 제목은 비슷하지만 문구 차이가 있어 확인을 권장합니다.")
    if "YOUTUBE_TITLE_FETCH_FAIL" in issues:
        messages.append("유튜브 실제 제목을 가져오지 못했습니다. 네트워크/비공개 영상 여부를 확인하세요.")
    if not messages and detail_notes:
        messages.append("실제 연결 영상은 같지만 URL 세부값이 다릅니다.")
    return " ".join(messages)


def summarize_link_result(qr_status: str, link_status: str, qr_link_status: str) -> str:
    if qr_status == "MISSING" and link_status == "MISSING":
        return "QR/하이퍼링크 없음"
    if qr_status == "MISSING":
        return "QR 없음"
    if link_status == "MISSING":
        return "하이퍼링크 없음"
    bad = {"MISMATCH", "MISSING", "MISSING_QR", "MISSING_LINK"}
    if qr_status in bad or link_status in bad or qr_link_status in {"MISMATCH", "MISSING_QR", "MISSING_LINK"}:
        return "불일치"
    if qr_status != "EXACT" or link_status != "EXACT" or qr_link_status != "EXACT":
        return "실질 일치 / 세부값 확인"
    return "일치"


def validate_video_cards(
    doc: fitz.Document,
    csv_df: pd.DataFrame,
    qrs: List[PdfQr],
    links: List[PdfLink],
    title_y_min: float,
    title_y_max: float,
    semantic_match_ok: bool = True,
    auto_title_region: bool = True,
    check_youtube_titles: bool = False,
    youtube_ok_threshold: float = 0.82,
    youtube_check_threshold: float = 0.62,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    df = csv_df.copy()
    for col in REQUIRED_COLUMNS:
        if col not in df.columns:
            raise ValueError(f"CSV에 '{col}' 컬럼이 없습니다. 현재 컬럼: {list(df.columns)}")

    df["페이지"] = pd.to_numeric(df["페이지"], errors="coerce").astype("Int64")
    df["카드번호"] = pd.to_numeric(df["카드번호"], errors="coerce").astype("Int64")
    df = df.dropna(subset=["페이지", "카드번호"]).copy()
    df["페이지"] = df["페이지"].astype(int)
    df["카드번호"] = df["카드번호"].astype(int)

    qrs_by_page = group_by_page([q for q in qrs if is_external_video_url(q.url)])
    links_by_page = group_by_page([l for l in links if l.url and is_external_video_url(l.url)])
    card_count_by_page = df.groupby("페이지")["카드번호"].max().to_dict()

    matched_qr_keys = set()
    matched_link_keys = set()
    results = []

    for _, r in df.sort_values(["페이지", "카드번호"]).iterrows():
        page_num = int(r["페이지"])
        card_no = int(r["카드번호"])
        expected_title = str(r["제목"]).strip()
        expected_url = str(r["유튜브링크"]).strip()

        page = doc[page_num - 1]
        page_qrs = qrs_by_page.get(page_num, [])
        page_links = links_by_page.get(page_num, [])
        qr = page_qrs[card_no - 1] if card_no - 1 < len(page_qrs) else None
        link = page_links[card_no - 1] if card_no - 1 < len(page_links) else None

        if qr:
            matched_qr_keys.add((qr.page, round(qr.x0, 2), round(qr.y0, 2), qr.url))
        if link:
            matched_link_keys.add((link.page, round(link.x0, 2), round(link.y0, 2), link.url))

        card_count = int(card_count_by_page.get(page_num, 4))
        x0, _, x1, _ = row_card_rect(page, card_count, card_no, link, qr)
        if auto_title_region:
            pdf_title_region, title_score = find_best_title_in_card(page, expected_title, x0, x1, title_y_min, title_y_max)
        else:
            pdf_title_region = extract_title_text_for_region(page, x0, x1, title_y_min, title_y_max)
            title_score = title_similarity(pdf_title_region, expected_title)

        title_region_match = normalize_text(expected_title) in normalize_text(pdf_title_region) or title_score >= 0.88
        title_page_match = page_has_text(page, expected_title)

        qr_url = qr.url if qr else ""
        link_url = link.url if link else ""
        qr_status, qr_note = compare_urls(expected_url, qr_url)
        link_status, link_note = compare_urls(expected_url, link_url)
        qr_link_status, qr_link_note = compare_two_actual_urls(qr_url, link_url)

        title_status = "OK" if title_region_match else ("PAGE_ONLY" if title_page_match else "MISMATCH")
        issues = []
        detail_notes = []

        if title_status == "MISMATCH":
            issues.append("TITLE_MISMATCH")
        elif title_status == "PAGE_ONLY":
            issues.append("TITLE_CHECK")
            detail_notes.append("제목은 페이지 전체에는 있으나 카드 영역에서 정확히 잡히지 않음")

        detail_issue_map = {
            "QR": (qr_status, qr_note),
            "LINK": (link_status, link_note),
            "QR_LINK": (qr_link_status, qr_link_note),
        }

        if qr_status == "MISSING":
            issues.append("MISSING_QR")
        elif qr_status == "MISMATCH":
            issues.append("QR_MISMATCH")
        elif qr_status != "EXACT":
            detail_notes.append(f"QR: {qr_note}")

        if link_status == "MISSING":
            issues.append("MISSING_LINK")
        elif link_status == "MISMATCH":
            issues.append("LINK_MISMATCH")
        elif link_status != "EXACT":
            detail_notes.append(f"LINK: {link_note}")

        if qr_link_status in {"MISMATCH", "MISSING_QR", "MISSING_LINK"}:
            if qr_url or link_url:
                issues.append("QR_LINK_MISMATCH")
        elif qr_link_status not in {"EXACT", "MISSING"}:
            detail_notes.append(f"QR↔LINK: {qr_link_note}")

        yt_title = ""
        yt_title_url = link_url or qr_url or expected_url
        yt_fetch_note = ""
        yt_status = "SKIP"
        yt_csv_score = 0.0
        yt_pdf_score = 0.0
        yt_note = ""
        if check_youtube_titles:
            yt_title, yt_fetch_note = fetch_youtube_title_cached(yt_title_url)
            if yt_fetch_note and not yt_title:
                parts = parse_url_parts(yt_title_url)
                if parts.get("type") == "video":
                    issues.append("YOUTUBE_TITLE_FETCH_FAIL")
                yt_status = "SKIP"
                yt_note = yt_fetch_note
            else:
                yt_status, yt_csv_score, yt_pdf_score, yt_note = judge_youtube_title(
                    yt_title,
                    expected_title,
                    pdf_title_region,
                    youtube_ok_threshold,
                    youtube_check_threshold,
                )
                if yt_status == "MISMATCH":
                    issues.append("YOUTUBE_TITLE_MISMATCH")
                elif yt_status == "CHECK":
                    issues.append("YOUTUBE_TITLE_CHECK")
                if yt_note:
                    detail_notes.append(f"YOUTUBE: {yt_note}")

        if issues:
            # Ambiguous title/details can be CHECK, actual mismatch/missing should be ISSUE.
            check_only = {"TITLE_CHECK", "YOUTUBE_TITLE_CHECK"}
            overall = "CHECK" if all(i in check_only for i in issues) else "ISSUE"
        else:
            semantic_ok_values = {"EXACT", "VIDEO_ID_MATCH", "PLAYLIST_MATCH", "CHANNEL_MATCH"}
            if semantic_match_ok:
                overall = "OK" if qr_status in semantic_ok_values and link_status in semantic_ok_values else "CHECK"
            else:
                overall = "OK" if qr_status == "EXACT" and link_status == "EXACT" and qr_link_status == "EXACT" else "CHECK"

        if overall == "OK" and detail_notes:
            # Keep OK, but note details such as index differences.
            pass

        link_summary = summarize_link_result(qr_status, link_status, qr_link_status)
        yt_display = "-"
        if check_youtube_titles:
            if yt_status == "OK":
                yt_display = "유사"
            elif yt_status == "CHECK":
                yt_display = "확인"
            elif yt_status == "MISMATCH":
                yt_display = "불일치"
            else:
                yt_display = "조회불가/생략"

        issue_text = ", ".join(dict.fromkeys(issues)) if issues else ""
        note_text = " / ".join(detail_notes)
        action_text = action_memo(issues, detail_notes)

        results.append(
            {
                "상태": overall,
                "문제유형": issue_text,
                "페이지": page_num,
                "카드번호": card_no,
                "제목": expected_title,
                "CSV 제목": expected_title,
                "PDF 제목": pdf_title_region,
                "제목검수": title_status,
                "제목유사도": title_score,
                "링크검수": link_summary,
                "CSV URL": expected_url,
                "QR URL": qr_url,
                "QR검수": qr_status,
                "하이퍼링크 URL": link_url,
                "하이퍼링크검수": link_status,
                "QR-하이퍼링크검수": qr_link_status,
                "유튜브 제목 검수": yt_display,
                "유튜브 실제 제목": yt_title,
                "유튜브 제목 조회 URL": yt_title_url if check_youtube_titles else "",
                "유튜브-CSV 유사도": yt_csv_score,
                "유튜브-PDF 유사도": yt_pdf_score,
                "비고": note_text,
                "조치 메모": action_text,
            }
        )

    extra_rows = []
    for qr in qrs:
        key = (qr.page, round(qr.x0, 2), round(qr.y0, 2), qr.url)
        if key not in matched_qr_keys:
            extra_rows.append(
                {
                    "구분": "EXTRA_QR",
                    "페이지": qr.page,
                    "x0": round(qr.x0, 1),
                    "y0": round(qr.y0, 1),
                    "URL/대상": qr.url,
                    "비고": "CSV 영상 카드 기준에 매칭되지 않은 QR",
                }
            )

    for link in links:
        key = (link.page, round(link.x0, 2), round(link.y0, 2), link.url)
        if key not in matched_link_keys:
            if link.kind == "internal":
                kind = "INTERNAL_LINK"
                target = f"page {link.target_page}" if link.target_page else "internal"
                note = "PDF 내부 페이지 이동 링크"
            else:
                kind = "EXTRA_LINK"
                target = link.url
                note = "CSV 영상 카드 기준에 매칭되지 않은 외부 링크"
            extra_rows.append(
                {
                    "구분": kind,
                    "페이지": link.page,
                    "x0": round(link.x0, 1),
                    "y0": round(link.y0, 1),
                    "URL/대상": target,
                    "비고": note,
                }
            )

    return pd.DataFrame(results), pd.DataFrame(extra_rows)


# -----------------------------
# Display helpers
# -----------------------------
def compact_view(video_df: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "상태", "페이지", "카드번호", "제목", "PDF 제목", "제목검수", "링크검수", "유튜브 제목 검수", "문제유형", "조치 메모",
    ]
    out = video_df[[c for c in cols if c in video_df.columns]].copy()
    out["카드"] = out.pop("카드번호") if "카드번호" in out.columns else ""
    ordered = ["상태", "페이지", "카드", "제목", "PDF 제목", "제목검수", "링크검수", "유튜브 제목 검수", "문제유형", "조치 메모"]
    return out[[c for c in ordered if c in out.columns]]


def filter_df(df: pd.DataFrame, status_filter: List[str], issue_filter: List[str], search_text: str, problem_only: bool) -> pd.DataFrame:
    out = df.copy()
    if problem_only:
        out = out[out["상태"].isin(["CHECK", "ISSUE"])]
    if status_filter:
        out = out[out["상태"].isin(status_filter)]
    if issue_filter:
        pattern = "|".join(re.escape(x) for x in issue_filter)
        out = out[out["문제유형"].fillna("").str.contains(pattern, case=False, regex=True)]
    if search_text.strip():
        search = search_text.strip().lower()
        searchable = out.fillna("").astype(str).agg(" ".join, axis=1).str.lower()
        out = out[searchable.str.contains(re.escape(search), regex=True)]
    return out


def status_help_text(video_df: pd.DataFrame) -> str:
    issue_count = int((video_df["상태"] == "ISSUE").sum())
    check_count = int((video_df["상태"] == "CHECK").sum())
    if issue_count:
        return f"수정이 필요한 ISSUE가 {issue_count}개 있습니다. 먼저 문제 항목 탭에서 확인하세요."
    if check_count:
        return f"치명적 오류는 없고 CHECK가 {check_count}개 있습니다. URL 세부값/제목 유사도를 확인하면 됩니다."
    return "모든 영상 카드가 OK입니다."


def render_page_preview(doc: fitz.Document, page_num: int, qrs: List[PdfQr], links: List[PdfLink], zoom: float = 1.0) -> Image.Image:
    page = doc[page_num - 1]
    matrix = fitz.Matrix(zoom, zoom)
    pix = page.get_pixmap(matrix=matrix, alpha=False)
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    draw = ImageDraw.Draw(img)

    # QR: red boxes, Links: blue boxes. RGB values are only for generated preview annotation.
    for qr in [q for q in qrs if q.page == page_num]:
        rect = [qr.x0 * zoom, qr.y0 * zoom, qr.x1 * zoom, qr.y1 * zoom]
        draw.rectangle(rect, outline=(220, 40, 40), width=4)
        draw.text((rect[0], max(0, rect[1] - 18)), "QR", fill=(220, 40, 40))

    for link in [l for l in links if l.page == page_num]:
        rect = [link.x0 * zoom, link.y0 * zoom, link.x1 * zoom, link.y1 * zoom]
        color = (30, 100, 220) if link.kind == "external" else (120, 120, 120)
        draw.rectangle(rect, outline=color, width=3)
        draw.text((rect[0], max(0, rect[1] - 18)), "LINK" if link.kind == "external" else "INTERNAL", fill=color)
    return img


# -----------------------------
# Excel report
# -----------------------------
def autosize_sheet(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 70))
        ws.column_dimensions[col_letter].width = max(max_len + 2, 10)


def write_dataframe(ws, df: pd.DataFrame):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c_idx, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.font = Font(bold=True, color="1F2A44")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    status_colors = {
        "OK": "E2F0D9",
        "CHECK": "FFF2CC",
        "ISSUE": "FCE4D6",
        "EXTRA_QR": "EADCF8",
        "EXTRA_LINK": "EADCF8",
        "INTERNAL_LINK": "E7E6E6",
    }

    for r_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
        status_value = str(row[0]) if row else ""
        row_fill = PatternFill("solid", fgColor=status_colors.get(status_value, "FFFFFF"))
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if c_idx == 1:
                cell.font = Font(bold=True)
            cell.fill = row_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize_sheet(ws)


def build_excel_report(video_df: pd.DataFrame, extra_df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "요약"

    total = len(video_df)
    ok = int((video_df["상태"] == "OK").sum()) if total else 0
    check = int((video_df["상태"] == "CHECK").sum()) if total else 0
    issue = int((video_df["상태"] == "ISSUE").sum()) if total else 0
    extra_count = len(extra_df)

    summary_rows = [
        ["항목", "개수"],
        ["CSV 기준 영상 카드", total],
        ["OK", ok],
        ["CHECK", check],
        ["ISSUE", issue],
        ["기타 QR/링크", extra_count],
    ]
    for row in summary_rows:
        ws_summary.append(row)
    for cell in ws_summary[1]:
        cell.font = Font(bold=True, color="1F2A44")
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    autosize_sheet(ws_summary)

    ws_compact = wb.create_sheet("보기용_요약")
    write_dataframe(ws_compact, compact_view(video_df))

    ws_problem = wb.create_sheet("문제항목만")
    problem_df = video_df[video_df["상태"].isin(["CHECK", "ISSUE"])].copy()
    write_dataframe(ws_problem, compact_view(problem_df) if not problem_df.empty else pd.DataFrame(columns=compact_view(video_df).columns))

    ws_video = wb.create_sheet("상세_전체데이터")
    write_dataframe(ws_video, video_df)

    ws_extra = wb.create_sheet("기타_QR_링크")
    if extra_df.empty:
        extra_df = pd.DataFrame(columns=["구분", "페이지", "x0", "y0", "URL/대상", "비고"])
    write_dataframe(ws_extra, extra_df)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# -----------------------------
# Main UI
# -----------------------------
def main():
    st.set_page_config(page_title="PDF QR 링크 자동 검수기", page_icon="✅", layout="wide")
    st.title("PDF QR · 하이퍼링크 자동 검수기")
    st.caption("CSV의 페이지/카드번호/제목/유튜브링크를 기준으로 PDF 안의 영상 제목, QR, 클릭 링크, 유튜브 실제 제목을 비교합니다.")

    with st.sidebar:
        st.header("검수 옵션")
        zoom = st.slider("QR 인식 해상도", min_value=1.0, max_value=4.0, value=2.0, step=0.5)
        st.caption("QR 인식 실패가 있으면 2.5 또는 3.0으로 올려보세요.")
        title_y_min = st.number_input("제목 탐색 Y 시작", min_value=0, max_value=1080, value=520, step=5)
        title_y_max = st.number_input("제목 탐색 Y 끝", min_value=0, max_value=1080, value=600, step=5)
        auto_title_region = st.checkbox("제목 위치 자동 보정", value=True)
        st.caption("카드 영역 안에서 CSV 제목과 가장 비슷한 텍스트를 찾습니다. 켜두는 것을 권장합니다.")
        semantic_match_ok = st.checkbox("영상 ID가 같으면 OK 처리", value=True)
        st.caption("체크 해제 시 index 등 URL 세부값이 다르면 CHECK로 표시됩니다.")
        st.divider()
        check_youtube_titles = st.checkbox("유튜브 실제 영상 제목까지 검수", value=False)
        youtube_ok_threshold = st.slider("유튜브 제목 OK 기준", 0.50, 1.00, 0.82, 0.01)
        youtube_check_threshold = st.slider("유튜브 제목 CHECK 기준", 0.30, 0.95, 0.62, 0.01)
        st.caption("유튜브 제목과 CSV/PDF 제목의 유사도 기준입니다. 기본값 권장.")

    pdf_file = st.file_uploader("PDF 파일 업로드", type=["pdf"])
    csv_file = st.file_uploader("CSV 기준 데이터 업로드", type=["csv"])

    if not pdf_file or not csv_file:
        st.info("PDF와 CSV를 모두 업로드한 뒤 검수를 실행해 주세요.")
        return

    if st.button("검수 실행", type="primary"):
        try:
            pdf_bytes = pdf_file.read()
            csv_bytes = csv_file.read()
            csv_df = read_csv_flexible(csv_bytes)

            missing = [c for c in REQUIRED_COLUMNS if c not in csv_df.columns]
            if missing:
                st.error(f"CSV 필수 컬럼이 없습니다: {missing}")
                st.stop()

            with st.spinner("PDF에서 QR과 하이퍼링크를 추출하는 중..."):
                doc = load_pdf(pdf_bytes)
                expected_counts = (
                    csv_df.groupby("페이지")["카드번호"].max().dropna().astype(int).to_dict()
                    if "페이지" in csv_df.columns and "카드번호" in csv_df.columns
                    else None
                )
                qrs = extract_qrs(doc, zoom=zoom, expected_counts=expected_counts)
                links = extract_links(doc)

            if check_youtube_titles:
                st.info("유튜브 실제 제목을 조회합니다. 영상 개수에 따라 시간이 조금 걸릴 수 있어요.")
                progress = st.progress(0)
                # validate_video_cards will fetch titles; progress is approximate.
                progress.progress(10)
            else:
                progress = None

            with st.spinner("CSV 기준으로 제목 / QR / 하이퍼링크 / 유튜브 제목을 비교하는 중..."):
                video_df, extra_df = validate_video_cards(
                    doc=doc,
                    csv_df=csv_df,
                    qrs=qrs,
                    links=links,
                    title_y_min=float(title_y_min),
                    title_y_max=float(title_y_max),
                    semantic_match_ok=semantic_match_ok,
                    auto_title_region=auto_title_region,
                    check_youtube_titles=check_youtube_titles,
                    youtube_ok_threshold=float(youtube_ok_threshold),
                    youtube_check_threshold=float(youtube_check_threshold),
                )
                if progress is not None:
                    progress.progress(100)
                    time.sleep(0.2)
                    progress.empty()

            st.session_state["result"] = {
                "video_df": video_df,
                "extra_df": extra_df,
                "pdf_bytes": pdf_bytes,
                "qrs": qrs,
                "links": links,
                "page_count": doc.page_count,
            }
            st.success("검수가 완료되었습니다.")
        except Exception as exc:
            st.exception(exc)
            return

    result = st.session_state.get("result")
    if not result:
        return

    video_df = result["video_df"]
    extra_df = result["extra_df"]
    ok = int((video_df["상태"] == "OK").sum())
    check = int((video_df["상태"] == "CHECK").sum())
    issue = int((video_df["상태"] == "ISSUE").sum())

    st.subheader("검수 요약")
    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("CSV 영상 카드", len(video_df))
    col2.metric("OK", ok)
    col3.metric("CHECK", check)
    col4.metric("ISSUE", issue)
    col5.metric("기타 QR/링크", len(extra_df))
    st.info(status_help_text(video_df))

    issue_counts = (
        video_df["문제유형"].fillna("").str.split(", ").explode().replace("", np.nan).dropna().value_counts()
    )
    if not issue_counts.empty:
        with st.expander("문제유형별 개수 보기", expanded=False):
            st.dataframe(issue_counts.rename_axis("문제유형").reset_index(name="개수"), hide_index=True, use_container_width=True)

    st.subheader("결과 보기")
    filter_cols = st.columns([1.2, 1.5, 1.5, 3])
    with filter_cols[0]:
        problem_only = st.checkbox("문제 항목만 보기", value=True)
    with filter_cols[1]:
        status_filter = st.multiselect("상태", options=["OK", "CHECK", "ISSUE"], default=[])
    with filter_cols[2]:
        all_issues = sorted({x for cell in video_df["문제유형"].fillna("") for x in cell.split(", ") if x})
        issue_filter = st.multiselect("문제유형", options=all_issues, default=[])
    with filter_cols[3]:
        search_text = st.text_input("검색", placeholder="제목, URL, 문제유형, 메모 검색")

    filtered = filter_df(video_df, status_filter, issue_filter, search_text, problem_only)
    compact_df = compact_view(filtered)

    tab1, tab2, tab3, tab4 = st.tabs(["보기용 요약", "상세 원자료", "기타 QR/링크", "페이지 미리보기"])

    with tab1:
        st.caption("URL은 숨기고 판단에 필요한 항목만 보여줍니다. 상세 URL은 '상세 원자료' 탭에서 확인하세요.")
        st.dataframe(
            compact_df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "상태": st.column_config.TextColumn("상태", width="small"),
                "페이지": st.column_config.NumberColumn("페이지", width="small"),
                "카드": st.column_config.NumberColumn("카드", width="small"),
                "제목": st.column_config.TextColumn("CSV 제목", width="medium"),
                "PDF 제목": st.column_config.TextColumn("PDF 제목", width="medium"),
                "조치 메모": st.column_config.TextColumn("조치 메모", width="large"),
            },
        )

        if not filtered.empty:
            st.markdown("#### 선택 항목 상세 확인")
            options = [f"p{int(r['페이지'])}-card{int(r['카드번호'])} | {r['제목']}" for _, r in filtered.iterrows()]
            selected = st.selectbox("상세를 볼 항목", options=options)
            sel_idx = options.index(selected)
            row = filtered.iloc[sel_idx]
            detail_cols = st.columns(3)
            detail_cols[0].markdown(f"**CSV 제목**  \n{row.get('CSV 제목','')}")
            detail_cols[1].markdown(f"**PDF 제목**  \n{row.get('PDF 제목','')}")
            detail_cols[2].markdown(f"**유튜브 실제 제목**  \n{row.get('유튜브 실제 제목','') or '-'}")
            st.markdown("**링크 상세**")
            st.code(
                f"CSV URL:\n{row.get('CSV URL','')}\n\nQR URL:\n{row.get('QR URL','')}\n\n하이퍼링크 URL:\n{row.get('하이퍼링크 URL','')}",
                language="text",
            )
            st.markdown(f"**조치 메모:** {row.get('조치 메모','')}")

    with tab2:
        st.caption("검수 원자료 전체입니다. URL 전체 비교가 필요할 때 사용하세요.")
        st.dataframe(filtered, use_container_width=True, hide_index=True)

    with tab3:
        st.caption("CSV 영상 카드 기준에 매칭되지 않은 QR, 외부 링크, PDF 내부 이동 링크입니다.")
        st.dataframe(extra_df, use_container_width=True, hide_index=True)

    with tab4:
        st.caption("빨간 박스는 QR, 파란 박스는 외부 하이퍼링크, 회색 박스는 내부 페이지 이동 링크입니다.")
        page_options = list(range(1, int(result["page_count"]) + 1))
        if not filtered.empty:
            default_page = int(filtered.iloc[0]["페이지"])
        else:
            default_page = page_options[0]
        page_num = st.selectbox("미리보기 페이지", options=page_options, index=page_options.index(default_page))
        preview_zoom = st.slider("미리보기 확대", 0.5, 2.0, 1.0, 0.25)
        doc = load_pdf(result["pdf_bytes"])
        img = render_page_preview(doc, int(page_num), result["qrs"], result["links"], zoom=float(preview_zoom))
        st.image(img, use_container_width=True)

    excel_bytes = build_excel_report(video_df, extra_df)
    st.download_button(
        label="엑셀 검수표 다운로드",
        data=excel_bytes,
        file_name="pdf_qr_link_validation_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    main()
